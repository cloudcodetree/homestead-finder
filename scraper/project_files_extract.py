"""Text-extraction worker for project_files.

Polls Supabase for `project_files` rows where `extracted_text IS NULL`,
downloads each blob from Storage, runs format-specific extraction,
and writes the result back. Idempotent: skips rows already extracted
unless --force is passed.

Format support:
  * .pdf  → pdfplumber (text layer; OCR for scanned PDFs is BACKLOG P2)
  * .docx → python-docx
  * .xlsx → openpyxl (every sheet → CSV-like text)
  * .txt / .md / .csv / .json → utf-8 read with surrogate-escape
  * images → SKIPPED for now; planned via Claude Vision in a
    future increment

Designed to run on a schedule (manual or cron) — not in-line on upload.
The frontend renders a "Processing…" state when extracted_text is NULL.

Usage:
    python -m project_files_extract
    python -m project_files_extract --force                 # re-extract everything
    python -m project_files_extract --file-id <uuid>        # one row
    python -m project_files_extract --limit 10              # cap per run
"""

from __future__ import annotations

import argparse
import hashlib
import io
import os
import sys
from typing import Any

import requests

import config
from logger import get_logger

log = get_logger("project_files_extract")

# ── Limits ─────────────────────────────────────────────────────────
# Hard cap on extracted_text size we'll write back. Most PDFs are
# fine; massive 500-page legal PDFs get truncated at the boundary.
# 200KB ≈ ~50K tokens which is enough for in-project AI context until
# the v2 chunks/pgvector path lands.
MAX_TEXT_BYTES = 200_000

# Bucket name must match frontend (api.ts: PROJECT_FILES_BUCKET).
BUCKET = "project-files"


def _supabase_headers(*, json_body: bool = True) -> dict[str, str]:
    """We read project_files across users (the worker is server-side,
    not user-scoped) — requires an RLS-bypassing key. Prefers the new
    `SUPABASE_SECRET_KEY` (API Keys 2.0); falls back to legacy
    `SUPABASE_SERVICE_ROLE_KEY` for projects still on JWT keys."""
    key = (
        os.environ.get("SUPABASE_SECRET_KEY")
        or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_ANON_KEY", "")
    )
    if not key:
        raise RuntimeError(
            "SUPABASE_SECRET_KEY (preferred) or SUPABASE_SERVICE_ROLE_KEY "
            "(legacy) required (we read across users)"
        )
    h = {"apikey": key, "Authorization": f"Bearer {key}"}
    if json_body:
        h["Content-Type"] = "application/json"
    return h


def _supabase_base() -> str:
    base = os.environ.get("SUPABASE_URL", "").rstrip("/")
    if not base:
        raise RuntimeError("SUPABASE_URL env var required")
    return base


# ── Extractors ─────────────────────────────────────────────────────
# Each returns plain text or "" on failure. Never raises — extraction
# is best-effort; an unsupported format leaves extracted_text NULL,
# the frontend handles that.


def _extract_pdf(content: bytes) -> str:
    """PDF text layer via pdfplumber (already a dep for tax-sale
    parsing). Page-by-page concatenation with form-feed separators."""
    try:
        import pdfplumber
    except ImportError:
        log.info("[extract] pdfplumber not available — skipping PDF")
        return ""
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            pages: list[str] = []
            for p in pdf.pages:
                txt = p.extract_text() or ""
                if txt:
                    pages.append(txt)
            return "\f".join(pages)
    except Exception as e:  # noqa: BLE001
        log.info(f"[extract] PDF parse failed: {type(e).__name__}: {e}")
        return ""


def _extract_docx(content: bytes) -> str:
    """DOCX via python-docx. Concatenates paragraphs with double-newline."""
    try:
        from docx import Document
    except ImportError:
        log.info("[extract] python-docx not available — skipping DOCX")
        return ""
    try:
        doc = Document(io.BytesIO(content))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:  # noqa: BLE001
        log.info(f"[extract] DOCX parse failed: {type(e).__name__}: {e}")
        return ""


def _extract_xlsx(content: bytes) -> str:
    """XLSX via openpyxl. Each sheet rendered as CSV-like rows; sheets
    separated by '## SheetName' headers so the AI can tell them apart."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.info("[extract] openpyxl not available — skipping XLSX")
        return ""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        out: list[str] = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            out.append(f"## {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                # Skip fully-blank rows
                if any(c.strip() for c in cells):
                    out.append(",".join(cells))
        return "\n".join(out)
    except Exception as e:  # noqa: BLE001
        log.info(f"[extract] XLSX parse failed: {type(e).__name__}: {e}")
        return ""


def _extract_text(content: bytes) -> str:
    """Plain-text formats — UTF-8 with surrogate-escape for safety."""
    try:
        return content.decode("utf-8", errors="surrogateescape")
    except Exception:  # noqa: BLE001
        return ""


def extract(content: bytes, content_type: str | None, filename: str) -> str:
    """Dispatch on (content_type, filename suffix). Returns empty
    string if format isn't supported."""
    ct = (content_type or "").lower()
    name = (filename or "").lower()
    if "pdf" in ct or name.endswith(".pdf"):
        return _extract_pdf(content)
    if "wordprocessingml" in ct or name.endswith(".docx"):
        return _extract_docx(content)
    if "spreadsheetml" in ct or name.endswith(".xlsx"):
        return _extract_xlsx(content)
    if (
        ct.startswith("text/")
        or name.endswith((".txt", ".md", ".csv", ".json"))
    ):
        return _extract_text(content)
    # Unsupported (images, .doc, .xls, etc.). Leave NULL so a future
    # extractor can pick it up.
    return ""


# ── Supabase wire-up ───────────────────────────────────────────────


def fetch_pending_files(
    *, force: bool, limit: int | None, file_id: str | None
) -> list[dict[str, Any]]:
    """Return rows that need extraction. force=True returns ALL rows;
    otherwise only those with extracted_text IS NULL."""
    url = (
        f"{_supabase_base()}/rest/v1/project_files"
        "?select=id,user_id,project_id,filename,content_type,"
        "storage_path,extracted_text,text_hash"
    )
    if file_id:
        url += f"&id=eq.{file_id}"
    elif not force:
        url += "&extracted_text=is.null"
    if limit:
        url += f"&limit={limit}"
    r = requests.get(url, headers=_supabase_headers(json_body=False), timeout=30)
    r.raise_for_status()
    return r.json()


def download_blob(storage_path: str) -> bytes | None:
    """Pull the file bytes from Storage (service role bypasses RLS)."""
    url = f"{_supabase_base()}/storage/v1/object/{BUCKET}/{storage_path}"
    r = requests.get(
        url, headers=_supabase_headers(json_body=False), timeout=60
    )
    if r.status_code != 200:
        log.info(
            f"[extract] download failed for {storage_path}: {r.status_code}"
        )
        return None
    return r.content


def update_extracted(file_id: str, text: str) -> bool:
    """Write extracted_text + text_hash back. Returns True on success."""
    truncated = text[:MAX_TEXT_BYTES]
    text_hash = hashlib.sha256(truncated.encode()).hexdigest()[:16]
    url = (
        f"{_supabase_base()}/rest/v1/project_files?id=eq.{file_id}"
    )
    headers = {**_supabase_headers(), "Prefer": "return=minimal"}
    body = {"extracted_text": truncated, "text_hash": text_hash}
    r = requests.patch(url, headers=headers, json=body, timeout=20)
    if r.status_code not in (200, 204):
        log.info(
            f"[extract] update {file_id} failed: {r.status_code} {r.text[:200]}"
        )
        return False
    return True


# ── Driver ─────────────────────────────────────────────────────────


def process(
    *, force: bool = False, limit: int | None = None, file_id: str | None = None
) -> int:
    """Main entry. Returns number of files successfully extracted."""
    try:
        files = fetch_pending_files(
            force=force, limit=limit, file_id=file_id
        )
    except Exception as e:  # noqa: BLE001
        log.info(f"[extract] could not fetch project_files: {e}")
        return 0

    if not files:
        log.info("[extract] nothing to do")
        return 0

    log.info(f"[extract] processing {len(files)} files")
    done = 0
    for row in files:
        fid = row["id"]
        path = row["storage_path"]
        ct = row.get("content_type")
        name = row.get("filename", "")
        log.info(f"[extract] {name} ({fid[:8]})")
        blob = download_blob(path)
        if blob is None:
            continue
        text = extract(blob, ct, name)
        if not text:
            log.info(f"[extract] {name}: unsupported format or empty extraction")
            continue
        if update_extracted(fid, text):
            done += 1
            log.info(f"[extract] {name}: {len(text)} chars extracted")
    return done


def main() -> None:
    ap = argparse.ArgumentParser(prog="project_files_extract")
    ap.add_argument(
        "--force", action="store_true",
        help="Re-extract files that already have extracted_text",
    )
    ap.add_argument(
        "--limit", type=int, default=None,
        help="Cap per run (default: process all pending)",
    )
    ap.add_argument(
        "--file-id", help="Process exactly one file by id",
    )
    args = ap.parse_args()
    n = process(force=args.force, limit=args.limit, file_id=args.file_id)
    log.info(f"[extract] done: {n} file(s) extracted")
    sys.exit(0 if n >= 0 else 1)


if __name__ == "__main__":
    main()
